"""
ポケモンチャンピオンズ 対面ダメージ計算エンジン

データソース:
  data/ポケモン図鑑.xlsx      ... 種族値・タイプ相性・技データ（日本語技名の一次情報はこちら）
  data/技使用率データ.JSON     ... 使用率・性格・努力値配分・持ち物・特性・技（英語名、月替わり）
  data/move_names_en_ja.json ... 技使用率データ.JSON の英語技名 → 日本語技名の対応表
  data/abilities_ja.json     ... 特性名 → 効果の対応表（build/extract_abilities.py で生成）

レギュレーションM-B シングル / レベル50固定 / 個体値31 / 努力値は「能力ポイント」表記
  1ポイント = 努力値8 / 1体あたり合計66ポイントまで / 1ステータス最大32ポイント
"""
import json
import os
from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, 'data', 'ポケモン図鑑.xlsx')
JSON_PATH = os.path.join(ROOT, 'data', '技使用率データ.JSON')
MOVE_NAME_JSON_PATH = os.path.join(ROOT, 'data', 'move_names_en_ja.json')
ABILITY_JSON_PATH = os.path.join(ROOT, 'data', 'abilities_ja.json')

# 技使用率データ.JSON が月替わりで差し替えられる前提の、既知の最終フォールバックシート名。
# 対応表(move_names_en_ja.json)に無い技だけがここを参照する。月が変わっても更新は必須ではない
# （このシート自体は2026年7月のスナップショットのまま残しておいてよい。それでも対応表拡充までの
#   橋渡しとして機能する）。シートが無くても（将来消しても）ビルドは通る。
LEGACY_MOVE_SHEET = '202607技使用率'


class PokemonNotFoundError(Exception):
    """使用率データの pokemon_id が図鑑（チャンピオンズ図鑑シート）に存在しない。"""


class RegionFormError(Exception):
    """region_form に対応する図鑑エントリが見つからない。"""

# ---------------------------------------------------------------- データ読み込み

WB = load_workbook(XLSX, data_only=True)

# タイプ相性表: EFF[(攻撃タイプ, 防御タイプ)] = 倍率
EFF = {}
_rows = list(WB['相性表'].iter_rows(values_only=True))
_deftypes = [c for c in _rows[1][2:] if c]
for _r in _rows[2:]:
    if not _r[1]:
        continue
    for _i, _dt in enumerate(_deftypes):
        if _r[2 + _i] is not None:
            EFF[(_r[1], _dt)] = float(_r[2 + _i])

# xlsx側の技名の誤りを読み込み時に直す。英語名をそのまま音写してしまっているもの。
# 本来は data/ポケモン図鑑.xlsx を直すのが筋だが、シートを跨いで揃える必要があるのでここで吸収する。
MOVE_NAME_FIX = {
    'スピリットブレイク': 'ソウルクラッシュ',   # Spirit Break の正式和名はソウルクラッシュ
}


def fix_move_name(name):
    return MOVE_NAME_FIX.get(name, name)


# 技データ: MOVES[技名] = {type, cat, power, acc, pri, effect}
MOVES = {}
for _r in WB['技データ'].iter_rows(min_row=2, values_only=True):
    if not _r[0]:
        continue
    MOVES[fix_move_name(_r[0])] = dict(type=_r[1], cat=_r[2],
                        power=float(_r[3]) if _r[3] else 0,
                        acc=float(_r[4]) if _r[4] else None,
                        pri=_r[5], effect=_r[6])

# 図鑑: DEX[ポケモン名] = {t1, t2, ab, base}   ※メガ形態は「メガ○○」で別エントリ
DEX = {}
BY_DEX_NO = {}   # 図鑑番号 -> [通常形態名, メガ形態名, ...]
for _r in WB['チャンピオンズ図鑑'].iter_rows(min_row=2, values_only=True):
    if not _r[1]:
        continue
    _bs = str(_r[5]).split('-')
    if len(_bs) != 6:
        continue
    DEX[_r[1]] = dict(t1=_r[2], t2=_r[3] or None, ab=_r[4],
                      base=[int(x) for x in _bs])
    try:
        _num = int(str(_r[0]).replace(':', '').strip())
        BY_DEX_NO.setdefault(_num, []).append(_r[1])
    except (ValueError, TypeError):
        pass

# 技名の英語→日本語対応表（一次情報）。技使用率データ.JSON の技は英語名で入っているので、
# 月が変わってもこれを差し替える必要はない。対応表に無い技だけ LEGACY_MOVE_SHEET にフォールバックする。
MOVE_NAME_EN_JA = json.load(open(MOVE_NAME_JSON_PATH, encoding='utf-8'))

# 特性名 → 効果。ダメージ計算そのものには使わず、「この特性を計算に入れなくてよいか」を
# 人が判断するための参照データ。generate.py の ABILITY_HANDLING と突き合わせて、
# 未分類の特性が使用率データに出てきたら警告する。
ABILITIES = json.load(open(ABILITY_JSON_PATH, encoding='utf-8'))

# 旧・使用率シート（月次、フォールバック専用）: LEGACY_USE[ポケモン名] = {rank, moves}
# moves は「日本語技名 (採用率%)」の文字列。対応表(MOVE_NAME_EN_JA)に無い技を、同じ月の
# このシートの並び順から拾うためだけに使う。無くてもビルドは通る。
LEGACY_USE = {}
if LEGACY_MOVE_SHEET in WB.sheetnames:
    for _r in WB[LEGACY_MOVE_SHEET].iter_rows(min_row=2, values_only=True):
        if _r[1]:
            _mv = []
            for _x in _r[3:13]:
                if not _x:
                    continue
                _nm = str(_x).split(' (')[0]
                _mv.append(str(_x).replace(_nm, fix_move_name(_nm), 1))
            LEGACY_USE[_r[1]] = dict(rank=_r[0], moves=_mv)

# 使用率JSON（英語名・230体）
USAGE = json.load(open(JSON_PATH, encoding='utf-8'))

# JSONの region_form を図鑑側の日本語表記に対応させるためのキーワード。
# 例: 'samurott-hisui' -> 'ヒスイ' -> 図鑑の 'ダイケンキ(ヒスイ)'
#     'rotom-wash'     -> 'ウォッシュ' -> 図鑑の 'ウォッシュロトム'
# これを見ずに図鑑番号の先頭を取ると、ヒスイダイケンキが通常ダイケンキの種族値で計算される。
REGION_KEYWORD = {
    'alola': 'アローラ', 'hisui': 'ヒスイ', 'galar': 'ガラル',
    'wash': 'ウォッシュ', 'heat': 'ヒート', 'frost': 'フロスト',
    'fan': 'スピン', 'mow': 'カット', 'eternal': 'えいえん',
    'female': '♀', 'dusk': 'たそがれ', 'midnight': 'まよなか', 'midday': 'まひる',
    'paldea-combat': 'パルデア単', 'paldea-blaze': 'パルデア炎', 'paldea-aqua': 'パルデア水',
}
REGION_ANY = ('アローラ', 'ヒスイ', 'ガラル', 'パルデア', 'ウォッシュ', 'ヒート',
              'フロスト', 'スピン', 'カット', 'えいえん', 'たそがれ', 'まよなか', 'まひる')


def resolve_form(entry, want_mega=False):
    """JSONの1エントリから、図鑑上の正しいポケモン名を返す。
    region_form を無視すると別形態の種族値で計算してしまうので必ずこれを通すこと。
    pokemon_id が図鑑に無ければ PokemonNotFoundError、region_form が指定されているのに
    対応する図鑑エントリが無ければ RegionFormError を投げる。どちらも黙って別形態の
    種族値で計算しないための安全弁（過去にヒスイダイケンキを通常種で計算したバグがある）。"""
    names = BY_DEX_NO.get(entry['pokemon_id'], [])
    if not names:
        raise PokemonNotFoundError(str(entry['pokemon_id']))
    rf = entry.get('region_form') or ''
    # 'paldea-combat-breed' のような複合キーは長いものから先に照合する
    keyword = None
    for k in sorted(REGION_KEYWORD, key=len, reverse=True):
        if k in rf:
            keyword = REGION_KEYWORD[k]
            break
    if keyword:
        cands = [n for n in names if keyword in n]
        if not cands:
            raise RegionFormError(f"pokemon_id={entry['pokemon_id']} region_form={rf!r}")
    else:
        cands = [n for n in names if not any(r in n for r in REGION_ANY)]
        # ♂♀で分かれている種は、region_formが無い側を♂とみなす
        if not cands:
            cands = names
        elif len(cands) > 1 and any('♂' in n for n in cands):
            cands = [n for n in cands if '♀' not in n]
    if not cands:
        cands = names
    megas = [n for n in cands if n.startswith('メガ')]
    plains = [n for n in cands if not n.startswith('メガ')]
    if want_mega and megas:
        return megas[0], megas
    return (plains[0] if plains else cands[0]), megas

# ---------------------------------------------------------------- 実数値の計算

NATURE = {
    'lonely': ('atk', 'def'), 'brave': ('atk', 'spe'), 'adamant': ('atk', 'spa'),
    'naughty': ('atk', 'spd'), 'bold': ('def', 'atk'), 'relaxed': ('def', 'spe'),
    'impish': ('def', 'spa'), 'lax': ('def', 'spd'), 'timid': ('spe', 'atk'),
    'hasty': ('spe', 'def'), 'jolly': ('spe', 'spa'), 'naive': ('spe', 'spd'),
    'modest': ('spa', 'atk'), 'mild': ('spa', 'def'), 'quiet': ('spa', 'spe'),
    'rash': ('spa', 'spd'), 'calm': ('spd', 'atk'), 'gentle': ('spd', 'def'),
    'sassy': ('spd', 'spe'), 'careful': ('spd', 'spa'),
}
IDX = {'hp': 0, 'atk': 1, 'def': 2, 'spa': 3, 'spd': 4, 'spe': 5}
NAT_JA = {
    'adamant': 'いじっぱり', 'jolly': 'ようき', 'timid': 'おくびょう', 'modest': 'ひかえめ',
    'bold': 'ずぶとい', 'impish': 'わんぱく', 'calm': 'おだやか', 'careful': 'しんちょう',
    'naive': 'むじゃき', 'hasty': 'せっかち', 'lonely': 'さみしがり', 'brave': 'ゆうかん',
    'naughty': 'やんちゃ', 'relaxed': 'のんき', 'lax': 'のうてんき', 'mild': 'おっとり',
    'quiet': 'れいせい', 'rash': 'うっかりや', 'gentle': 'おとなしい', 'sassy': 'なまいき',
}


def stats(base, ev_points, nature=None):
    """種族値と能力ポイント(6要素)から実数値を返す。レベル50・個体値31固定。"""
    out = []
    for i in range(6):
        ev = ev_points[i] * 8
        if i == 0:
            out.append(int((2 * base[0] + 31 + ev // 4) * 50 // 100) + 60)
        else:
            out.append(int((2 * base[i] + 31 + ev // 4) * 50 // 100) + 5)
    if nature in NATURE:
        up, dn = NATURE[nature]
        out[IDX[up]] = int(out[IDX[up]] * 1.1)
        out[IDX[dn]] = int(out[IDX[dn]] * 0.9)
    return out


# ---------------------------------------------------------------- タイプ相性と特性

def eff(move_type, t1, t2=None):
    """タイプ相性倍率。t2 は None 可。"""
    e = EFF[(move_type, t1)]
    if t2:
        e *= EFF[(move_type, t2)]
    return e


# 防御側の特性による軽減・無効（日本語名と英語名の両方を受け付ける）
IMMUNE_JA = {'ふゆう': 'じめん', 'もらいび': 'ほのお', 'ちょすい': 'みず', 'よびみず': 'みず',
             'かんそうはだ': 'みず', 'ちくでん': 'でんき', 'ひらいしん': 'でんき',
             'でんきエンジン': 'でんき', 'そうしょく': 'くさ'}
IMMUNE_EN = {'levitate': 'じめん', 'flash-fire': 'ほのお', 'water-absorb': 'みず',
             'storm-drain': 'みず', 'dry-skin': 'みず', 'volt-absorb': 'でんき',
             'lightning-rod': 'でんき', 'motor-drive': 'でんき', 'sap-sipper': 'くさ',
             'earth-eater': 'じめん'}
HALF_JA = {'あついしぼう': ('ほのお', 'こおり'), 'たいねつ': ('ほのお',), 'すいほう': ('ほのお',)}
HALF_EN = {'thick-fat': ('ほのお', 'こおり'), 'heatproof': ('ほのお',), 'water-bubble': ('ほのお',)}
# ability_mod が返す特性名を表示用の日本語に揃える
ABILITY_DISPLAY = {
    'levitate': 'ふゆう', 'flash-fire': 'もらいび', 'water-absorb': 'ちょすい',
    'storm-drain': 'よびみず', 'dry-skin': 'かんそうはだ', 'volt-absorb': 'ちくでん',
    'lightning-rod': 'ひらいしん', 'motor-drive': 'でんきエンジン', 'sap-sipper': 'そうしょく',
    'earth-eater': '土食い', 'thick-fat': 'あついしぼう', 'heatproof': 'たいねつ',
    'water-bubble': 'すいほう',
}
SOUND = {'ハイパーボイス', 'うたかたのアリア', 'ばくおんぱ', 'いびき', 'エコーボイス', 'りんしょう'}


def ability_mod(ability, move_type, mold_breaker=False, hp_full=True, is_sound=False):
    """防御側特性による倍率と、発動した特性名を返す。
    mold_breaker=True（かたやぶり）なら防御特性を全て無視する。
    ばけのかわは倍率ではなく「1回無効」なのでここでは 1.0 を返し、呼び出し側でターン数に加算する。
    """
    ab = ability or ''
    if mold_breaker:
        return 1.0, ''
    for table in (IMMUNE_JA, IMMUNE_EN):
        for name, typ in table.items():
            if name in ab and typ == move_type:
                return 0.0, ABILITY_DISPLAY.get(name, name)
    for table in (HALF_JA, HALF_EN):
        for name, types in table.items():
            if name in ab and move_type in types:
                return 0.5, ABILITY_DISPLAY.get(name, name)
    if ('マルチスケイル' in ab or 'multiscale' in ab) and hp_full:
        return 0.5, 'マルチスケイル'
    if ('ぼうおん' in ab or 'soundproof' in ab) and is_sound:
        return 0.0, 'ぼうおん'
    if any(k in ab for k in ('ハードロック', 'フィルター', 'solid-rock', 'filter', 'prism-armor')):
        return 0.75, 'ハードロック'   # ※効果抜群のときのみ有効。呼び出し側で判定すること
    if 'ばけのかわ' in ab or 'disguise' in ab:
        return 1.0, 'ばけのかわ'
    return 1.0, ''


# ---------------------------------------------------------------- ダメージ計算

def damage(power, attack, defense, stab=1.0, type_eff=1.0, extra=1.0):
    """レベル50固定のダメージ計算。(最低乱数, 最高乱数) を返す。
    stab   : タイプ一致補正（通常1.5 / てきおうりょく2.0）
    type_eff: タイプ相性 × 防御特性の倍率
    extra  : その他の乗算補正（いのちのたま1.3 / きれあじ1.5 / フェアリースキン1.2 など）
    丸めは 基礎 → 乱数 → 一致 → 相性 → その他 の順に切り捨てる。
    """
    base = int(int(2 * 50 / 5 + 2) * power * attack / defense / 50) + 2

    def roll(r):
        x = int(base * r)
        x = int(x * stab)
        x = int(x * type_eff)
        x = int(x * extra)
        return max(1, x)

    return roll(0.85), roll(1.0)


def verdict(lo, hi, hp):
    """最低乱数・最高乱数・相手HPから確定何発かを返す。"""
    if lo >= hp:
        return '確1'
    if hi >= hp:
        return '乱1'
    if lo * 2 >= hp:
        return '確2'
    if hi * 2 >= hp:
        return '乱2'
    if lo * 3 >= hp:
        return '確3'
    return '4発+'


VERDICT_RANK = {'確1': 5, '乱1': 4, '確2': 3, '乱2': 2, '確3': 1, '4発+': 0}
