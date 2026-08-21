#!/usr/bin/env python3
"""
data/ポケモン図鑑.xlsx を CSV に書き出す（一度きりの移行用）。

    pip install openpyxl
    python build/migrate_to_csv.py

移行後は engine.py が CSV を読むので、通常のビルドに openpyxl も xlsx も要らない。
xlsx はリポジトリに残すが、コードは読まない。新しいポケモンや技を足すときは CSV を直接編集する。

特性の分割について:
  xlsx の特性欄は複数の特性が区切り無しで連結されている（'しんりょくようりょくそ'）。
  これだと分割できず部分一致でしか判定できないので、CSV では '/' 区切りにする。
  分割は既知の特性名から最長一致で行い、**つなぎ直して元の文字列と一致すること**を
  必ず確認する。一致しなければ分割せず1語として出し、警告する（情報は落とさない）。
"""
import csv
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, 'data', 'ポケモン図鑑.xlsx')
DATA = os.path.join(ROOT, 'data')

# 特性一覧（data/abilities_ja.json）に載っていない特性。
# チャンピオンズ独自のメガや新しめの特性で、Wikiの一覧に無いもの。
# ここに無いと連結を切れないので、図鑑に新しい特性を足したらここにも足す。
EXTRA_ABILITIES = [
    # 単独で使われているもの
    'メガソーラー', 'おやこあい', 'スカイスキン', 'ドラゴンスキン', 'フリーズスキン',
    'とびだすなかみ', 'とびだすハバネロ', 'かんつうドリル', 'うなぎのぼり', 'てんきや',
    'ミイラ', 'ぎたい', 'ほのおのたてがみ', 'フェアリーオーラ', 'さまようたましい',
    'マイティチェンジ', 'バトルスイッチ',
    # data/特性の効果.pdf（育成考察Wiki）に項目が無い新しめの特性
    'でんきにかえる', 'どくげしょう', 'そうだいしょう',
    # 連結の構成要素として要るもの
    'かんろなミツ', 'ねんちゃく', 'バリアフリー', 'テイルアーマー', 'はんすう',
    'いかりのつぼ', 'びんじょう', 'おもてなし', 'アイスボディ', 'ちどりあし',
    'そうしょく', 'さいせいりょく', 'おみとおし', 'かそく', 'たいねつ',
    'いかく', 'ゆきふらし',
]


def load_vocab():
    import json
    vocab = set(EXTRA_ABILITIES)
    path = os.path.join(DATA, 'abilities_ja.json')
    if os.path.exists(path):
        vocab |= set(json.load(open(path, encoding='utf-8')))
    return {v for v in vocab if v}


def split_abilities(text, vocab):
    """連結された特性名を最長一致で切る。切れなければ None。"""
    out, i = [], 0
    while i < len(text):
        cands = [w for w in vocab if text.startswith(w, i)]
        if not cands:
            return None
        w = max(cands, key=len)
        out.append(w)
        i += len(w)
    return out if ''.join(out) == text else None


def write_csv(name, header, rows):
    path = os.path.join(DATA, name)
    with open(path, 'w', encoding='utf-8', newline='') as f:
        w = csv.writer(f, lineterminator='\n')
        w.writerow(header)
        w.writerows(rows)
    print(f'  {name:16} {len(rows):5} 行')
    return path


def main():
    from openpyxl import load_workbook
    wb = load_workbook(XLSX, data_only=True)
    vocab = load_vocab()

    # ---------------- 図鑑
    unsplit = []
    dex_rows = []
    for r in wb['チャンピオンズ図鑑'].iter_rows(min_row=2, values_only=True):
        if not r[1]:
            continue
        base = str(r[5]).split('-')
        if len(base) != 6:
            continue
        no = str(r[0]).replace(':', '').strip()
        ab = r[4] or ''
        parts = split_abilities(ab, vocab) if ab else []
        if ab and parts is None:
            unsplit.append((r[1], ab))
            parts = [ab]
        dex_rows.append([no, r[1], r[2], r[3] or '', '/'.join(parts)] + [int(x) for x in base])
    write_csv('dex.csv', ['no', 'name', 'type1', 'type2', 'abilities',
                          'hp', 'atk', 'def', 'spa', 'spd', 'spe'], dex_rows)

    # ---------------- 技
    move_rows = []
    for r in wb['技データ'].iter_rows(min_row=2, values_only=True):
        if not r[0]:
            continue
        move_rows.append([r[0], r[1], r[2],
                          '' if r[3] is None else r[3],
                          '' if r[4] is None else r[4],
                          '' if r[5] is None else r[5],
                          r[6] or '',
                          r[8] or ''])   # ランク変化。積み技の判定に使う
    write_csv('moves.csv', ['name', 'type', 'category', 'power', 'accuracy',
                            'priority', 'effect', 'rank_change'], move_rows)

    # ---------------- タイプ相性
    rows = list(wb['相性表'].iter_rows(values_only=True))
    deftypes = [c for c in rows[1][2:] if c]
    chart = []
    for r in rows[2:]:
        if not r[1]:
            continue
        chart.append([r[1]] + ['' if r[2 + i] is None else r[2 + i]
                               for i in range(len(deftypes))])
    write_csv('type_chart.csv', ['attack'] + deftypes, chart)

    if unsplit:
        print()
        print('警告: 特性を分割できませんでした（1語として出しています）。')
        print('      EXTRA_ABILITIES に構成要素を足すと切れるようになります:')
        for name, ab in unsplit:
            print(f'  {name}: {ab}')
    else:
        print()
        print('特性はすべて分割できました。')


if __name__ == '__main__':
    main()
